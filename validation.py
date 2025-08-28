"""
Validation and reporting functions for comparing ToC and parsed sections.
"""
import json
import openpyxl
import logging
from typing import List, Dict, Any, Optional, Tuple, Set
from models import Section, ValidationResult
from openpyxl.styles import Font, PatternFill, Alignment

# Configure logging
logging.basicConfig(level=logging.INFO, 
                   format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DocumentValidator:
    """Class to handle validation between ToC and parsed sections."""
    
    def __init__(self):
        """Initialize the validator."""
        self.result = ValidationResult()
        
    def _load_jsonl(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Load JSONL file into a list of dictionaries.
        
        Args:
            file_path: Path to the JSONL file
            
        Returns:
            List of dictionaries
        """
        try:
            with open(file_path, encoding="utf-8") as f:
                return [json.loads(line) for line in f if line.strip()]
        except Exception as e:
            logger.error(f"Error loading {file_path}: {str(e)}")
            return []
    
    def _convert_to_sections(self, entries: List[Dict[str, Any]]) -> List[Section]:
        """
        Convert dictionary entries to Section objects.
        
        Args:
            entries: List of dictionary entries
            
        Returns:
            List of Section objects
        """
        sections = []
        for entry in entries:
            try:
                section = Section(
                    doc_title=entry.get('doc_title', ''),
                    section_id=entry.get('section_id', ''),
                    title=entry.get('title', ''),
                    page=entry.get('page', 0),
                    level=entry.get('level', 0),
                    parent_id=entry.get('parent_id'),
                    full_path=entry.get('full_path'),
                    text_content=entry.get('text_content'),
                    tags=entry.get('tags', [])
                )
                sections.append(section)
            except Exception as e:
                logger.warning(f"Error converting entry to Section: {str(e)}")
        return sections
    
    def _find_missing_sections(self, toc_sections: List[Section], 
                              parsed_sections: List[Section]) -> List[Section]:
        """Find sections in ToC that are missing from parsed sections."""
        toc_ids = {section.section_id for section in toc_sections}
        parsed_ids = {section.section_id for section in parsed_sections}
        missing_ids = toc_ids - parsed_ids
        
        return [section for section in toc_sections if section.section_id in missing_ids]
    
    def _find_extra_sections(self, toc_sections: List[Section], 
                            parsed_sections: List[Section]) -> List[Section]:
        """Find sections in parsed sections that are not in ToC."""
        toc_ids = {section.section_id for section in toc_sections}
        parsed_ids = {section.section_id for section in parsed_sections}
        extra_ids = parsed_ids - toc_ids
        
        return [section for section in parsed_sections if section.section_id in extra_ids]
    
    def _find_order_mismatches(self, toc_sections: List[Section], 
                              parsed_sections: List[Section]) -> List[Dict[str, Any]]:
        """Find order mismatches between ToC and parsed sections."""
        mismatches = []
        
        # Create mapping of section_id to index
        toc_id_to_idx = {section.section_id: i for i, section in enumerate(toc_sections)}
        parsed_id_to_idx = {section.section_id: i for i, section in enumerate(parsed_sections)}
        
        # Find common section IDs
        common_ids = set(toc_id_to_idx.keys()) & set(parsed_id_to_idx.keys())
        
        # Check if the order differs
        for id1 in common_ids:
            for id2 in common_ids:
                if id1 != id2:
                    toc_order = toc_id_to_idx[id1] < toc_id_to_idx[id2]
                    parsed_order = parsed_id_to_idx[id1] < parsed_id_to_idx[id2]
                    
                    if toc_order != parsed_order:
                        mismatches.append({
                            "toc_id": id1,
                            "parsed_id": id2,
                            "comment": f"Order mismatch: {id1} {'before' if toc_order else 'after'} {id2} in ToC, "
                                     f"but {'after' if toc_order else 'before'} in parsed"
                        })
        
        return mismatches
    
    def _calculate_coverage(self, toc_sections: List[Section], 
                           parsed_sections: List[Section]) -> float:
        """Calculate content coverage percentage."""
        if not toc_sections:
            return 0.0
            
        common_count = len(set(s.section_id for s in toc_sections) & 
                           set(s.section_id for s in parsed_sections))
        return (common_count / len(toc_sections)) * 100.0
    
    def validate(self, toc_jsonl: str, spec_jsonl: str) -> ValidationResult:
        """
        Validate ToC against parsed sections.
        
        Args:
            toc_jsonl: Path to ToC JSONL file
            spec_jsonl: Path to parsed sections JSONL file
            
        Returns:
            ValidationResult object
        """
        # Load data
        toc_data = self._load_jsonl(toc_jsonl)
        spec_data = self._load_jsonl(spec_jsonl)
        
        # Convert to Section objects
        toc_sections = self._convert_to_sections(toc_data)
        parsed_sections = self._convert_to_sections(spec_data)
        
        # Find issues
        missing = self._find_missing_sections(toc_sections, parsed_sections)
        extra = self._find_extra_sections(toc_sections, parsed_sections)
        mismatches = self._find_order_mismatches(toc_sections, parsed_sections)
        coverage = self._calculate_coverage(toc_sections, parsed_sections)
        
        # Create result
        self.result = ValidationResult(
            missing_in_parsed=missing,
            extra_in_parsed=extra,
            order_mismatches=mismatches,
            toc_count=len(toc_sections),
            parsed_count=len(parsed_sections),
            coverage_percentage=coverage
        )
        
        return self.result
    
    def generate_report(self, report_xlsx: str) -> None:
        """
        Generate Excel report from validation results.
        
        Args:
            report_xlsx: Path to output Excel file
        """
        if not self.result:
            logger.warning("No validation results to report.")
            return
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Validation"
        
        # Add header with formatting
        header = ["Type", "Section ID", "Title", "Page", "Comment"]
        ws.append(header)
        
        # Format header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # Add missing sections
        for section in self.result.missing_in_parsed:
            ws.append([
                "Missing in Parsed",
                section.section_id,
                section.title,
                section.page,
                "In ToC, not found in parsed content"
            ])
            
        # Add extra sections
        for section in self.result.extra_in_parsed:
            ws.append([
                "Extra in Parsed",
                section.section_id,
                section.title,
                section.page,
                "In parsed content, not found in ToC"
            ])
            
        # Add order mismatches
        for mismatch in self.result.order_mismatches:
            ws.append([
                "Order Mismatch",
                f"{mismatch['toc_id']} / {mismatch['parsed_id']}",
                "",
                "",
                mismatch["comment"]
            ])
            
        # Add summary
        ws.append([])
        ws.append(["Summary", "", "", "", ""])
        ws.append(["Total in ToC", self.result.toc_count, "", "", ""])
        ws.append(["Total in Parsed", self.result.parsed_count, "", "", ""])
        ws.append(["Coverage", f"{self.result.coverage_percentage:.2f}%", "", "", ""])
        
        # Format summary
        summary_row = len(ws["A"]) - 4
        summary_font = Font(bold=True)
        summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        for row in range(summary_row, summary_row + 5):
            for cell in ws[row]:
                cell.font = summary_font
                cell.fill = summary_fill
                
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                    
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # Save workbook
        try:
            wb.save(report_xlsx)
            logger.info(f"Validation report saved to {report_xlsx}")
        except Exception as e:
            logger.error(f"Error saving report to {report_xlsx}: {str(e)}")

# For backward compatibility
def validate_and_report(toc_jsonl: str, spec_jsonl: str, report_xlsx: str) -> None:
    """
    Compares ToC and parsed sections, outputs XLS report.
    """
    validator = DocumentValidator()
    validator.validate(toc_jsonl, spec_jsonl)
    validator.generate_report(report_xlsx)
